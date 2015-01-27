"""
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000041 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000041">
        <rdf:type rdf:resource="&obo;MEBDO_0000042"/>
        <rdfs:label xml:lang="en">95</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000043 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000043">
        <rdf:type rdf:resource="&obo;MEBDO_0000031"/>
        <rdfs:label xml:lang="en">37467.9578947368</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000044 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000044">
        <rdf:type rdf:resource="&obo;MEBDO_0000045"/>
        <rdfs:label xml:lang="en">5525.67368421053</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000046 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000046">
        <rdf:type rdf:resource="&obo;MEBDO_0000047"/>
        <rdfs:label xml:lang="en">4485.87368421053</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000050 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000050">
        <rdf:type rdf:resource="&obo;MEBDO_3740039"/>
        <rdfs:label xml:lang="en">service_029_1</rdfs:label>
        <obo:MEBDO_0000417 rdf:resource="&obo;MEBDO_0000041"/> 
        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_0000043"/>
        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_0000044"/>
        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_0000046"/>
        <obo:MEBDO_0000026 rdf:resource="&obo;MEBDO_8010001"/>
    </owl:NamedIndividual>
 
 """
 
##hospital id will use 8+their id in the table. (ignor the reference region for the moment)
## assign column I(MEBDO_0000417) J K L (MEBDO_0000414) first, then add hospital individual (MEBDO_0000026)
## id starts from 7000000



from openpyxl import load_workbook

wb = load_workbook(filename = 'Medicare_Provider_Charge_Inpatient_DRG100_FY2012.xlsx')
ws = wb.get_sheet_by_name( name = 'Top_100_drg')

varId = 7000000

for i in range(8, 509):
    varId = 1 + varId 
    
    MSDRG = ws.cell('A'+ str(i)).value
    MSDGR = str(MSDRG)
    msDRGCode = MSDRG.split("-",2)[0]
    n = len(msDRGCode.strip())
    string_val = "0" * (4-n)
    MSDRGId = 'MEBDO_374' + string_val + msDRGCode.strip()
    
    hospitalId = ws.cell('B'+ str(i)).value
    hospitalId = str(hospitalId)
    hospitalId = '8' + str(hospitalId.strip())
    
    dischargeCount = ws.cell('I'+ str(i)).value
    charge = ws.cell('J'+ str(i)).value
    totalpayment = ws.cell('K'+ str(i)).value
    medicarepayment = ws.cell('L'+ str(i)).value
    
    updateFile = open('mebdo_DGR_payment_individual.owl','a')
    
    dischargeCountId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + dischargeCountId +' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_' + dischargeCountId +'">\n        <rdf:type rdf:resource="&obo;MEBDO_0000042"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ str(dischargeCount) +'</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')
    
    varId = varId + 1
    chargeId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_'+ chargeId +' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_'+ chargeId +'">\n        <rdf:type rdf:resource="&obo;MEBDO_0000031"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ str(charge) + '</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')
    
    varId = varId + 1
    totalpaymentId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + totalpaymentId +' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_' + totalpaymentId + '">\n        <rdf:type rdf:resource="&obo;MEBDO_0000045"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">' + str(totalpayment) + '</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')
    
    varId = varId + 1
    medicarepaymentId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + medicarepaymentId + ' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_'+ medicarepaymentId + '">\n        <rdf:type rdf:resource="&obo;MEBDO_0000047"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">' + str(medicarepayment) + '</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')
    
    varId = varId + 1
    inpatientServiceId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_'+ inpatientServiceId + ' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_' + inpatientServiceId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="&obo;' + MSDRGId + '"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">inpatient service ' + inpatientServiceId +'</rdfs:label>\n')
    updateFile.write('        <obo:MEBDO_0000417 rdf:resource="&obo;MEBDO_'+dischargeCountId+'"/>\n') 
    updateFile.write('        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_'+ chargeId + '"/>\n')
    updateFile.write('        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_'+  totalpaymentId + '"/>\n')
    updateFile.write('        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_' + medicarepaymentId + '"/>\n')
    updateFile.write('        <obo:MEBDO_0000026 rdf:resource="&obo;MEBDO_' + hospitalId +'"/>\n    </owl:NamedIndividual>\n\n\n')
    
    print i
    updateFile.close()
    


    
    
    

