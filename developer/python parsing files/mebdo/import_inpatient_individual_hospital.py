""" outpatient individual hospital
    <!-- http://purl.obolibrary.org/obo/MEBDO_0000408 -->

    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_0000408">
        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000001"/>
        <rdfs:label xml:lang="en">CENTRAL PENINSULA GENERAL HOSPITAL</rdfs:label>
        <obo:MEBDO_0000407 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000409"/>
        <obo:BFO_0000171 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000410"/>
        <obo:BFO_0000171 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000411"/>
        <obo:BFO_0000171 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000412"/>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000409 -->

    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_0000409">
        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000406"/>
        <rdfs:label xml:lang="en">250 HOSPITAL PLACE</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000410 -->

    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_0000410">
        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000010"/>
        <rdfs:label xml:lang="en">SOLDOTNA</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000411 -->

    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_0000411">
        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000011"/>
        <rdfs:label xml:lang="en">AK</rdfs:label>
    </owl:NamedIndividual>
    
    
    <!-- http://purl.obolibrary.org/obo/MEBDO_0000412 -->

    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_0000412">
        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000012"/>
        <rdfs:label xml:lang="en">99669</rdfs:label>
    </owl:NamedIndividual>
    


"""

## hospital id will use 8+their id in the table. (ignor the reference region for the moment)
## asign the column D (type:MEBDO_0000406), E(type:MEBDO_0000010), F(type: MEBDO_0000011), G(type:MEBDO_0000012) first, then assign the hospital (column C)
## id starts from 4000
## manually created head for the import file to declare the class and properties


from openpyxl import load_workbook

objectpropertyList = ['MEBDO_0000407','BFO_0000171']
classList = ['MEBDO_0000406','MEBDO_0000010','MEBDO_0000011','MEBDO_0000012']

fileName = 'mebdo_DRG_hospital_individual_import.owl'
updateFile =  open('mebdo_DRG_hospital_individual_import.owl','a')
updateFile.write('<?xml version="1.0"?>\n')
updateFile.write('<rdf:RDF xmlns="http://purl.obolirary.org/obo/mebdo/' + fileName +'#"\n')
updateFile.write('     xml:base="http://purl.obolirary.org/obo/mebdo/' + fileName + '"\n')
updateFile.write('     xmlns:obo="http://purl.obolibrary.org/obo/"\n     xmlns:rdfs="http://www.w3.org/2000/01/rdf-schema#"\n     xmlns:owl="http://www.w3.org/2002/07/owl#"\n     xmlns:xsd="http://www.w3.org/2001/XMLSchema#"\n     xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#">\n')
updateFile.write('    <owl:Ontology rdf:about="http://purl.obolirary.org/obo/mebdo/'+ fileName +'"/>\n\n\n')
updateFile.write('    <!-- \n    ///////////////////////////////////////////////////////////////////////////////////////\n    //\n    // Object Properties\n    //\n    ///////////////////////////////////////////////////////////////////////////////////////\n     -->\n\n')

m=len(objectpropertyList)
for i in range(m):
    objectProperty=objectpropertyList[i]
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/'+ objectProperty+ '  -->\n\n')
    updateFile.write('    <owl:ObjectProperty rdf:about="http://purl.obolibrary.org/obo/' + objectProperty + '"/>\n\n\n') 

updateFile.write('    <!-- \n    ///////////////////////////////////////////////////////////////////////////////////////\n    //\n    // Classes\n    //\n    ///////////////////////////////////////////////////////////////////////////////////////\n     -->\n\n\n')

m=len(classList)
for i in range(m):
    classname=classList[i]
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/' + classname + ' -->\n\n')
    updateFile.write('    <owl:Class rdf:about="http://purl.obolibrary.org/obo/' + classname + '"/>\n\n')
    

updateFile.write('      <!-- \n   ///////////////////////////////////////////////////////////////////////////////////////\n    //\n    // Individuals\n    //\n    ///////////////////////////////////////////////////////////////////////////////////////\n     -->\n\n\n')


wb = load_workbook(filename = 'Medicare_Provider_Charge_Inpatient_DRG100_FY2012.xlsx')
ws = wb.get_sheet_by_name( name = 'Top_100_drg')
varId = 4000 


for i in range(8, 509):
    varId = 1 + varId 
    hospitalId = ws.cell('B'+ str(i)).value
    hospital = ws.cell('C'+ str(i)).value
    streetAddress = ws.cell('D'+ str(i)).value
    city = ws.cell('E'+ str(i)).value
    state = ws.cell('F'+ str(i)).value
    zip = ws.cell('G'+ str(i)).value
    
    updateFile = open('mebdo_DRG_hospital_individual_import.owl','a')
    
    streetId = str(varId)
    n = len(streetId)
    streetId7digits = "0" * (7-n) + streetId 
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + streetId7digits)
    updateFile.write(' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_'+ streetId7digits + '">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000406"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ streetAddress )
    updateFile.write('</rdfs:label>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n')
    
    varId = varId + 1
    cityId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + cityId)
    updateFile.write(' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_'+ cityId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000010"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ city )
    updateFile.write('</rdfs:label\n>')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    
    varId = varId + 1
    stateId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + str(varId))
    updateFile.write(' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_'+ str(varId) + '">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000011"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ state )
    updateFile.write('</rdfs:label>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    varId = varId + 1
    zipId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + zipId)
    updateFile.write(' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_'+ zipId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000012"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ str(zip) )
    updateFile.write('</rdfs:label>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    
    hospitalId = '8' + str(hospitalId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_'+ hospitalId +' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="http://purl.obolibrary.org/obo/MEBDO_'+ hospitalId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000001"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ hospital +'</rdfs:label>\n')
    updateFile.write('        <obo:MEBDO_0000407 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_'+ streetId7digits + '"/>\n')
    updateFile.write('        <obo:BFO_0000171 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_'+ cityId +'"/>\n')
    updateFile.write('        <obo:BFO_0000171 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_'+ stateId +'"/>\n')
    updateFile.write('        <obo:BFO_0000171 rdf:resource="http://purl.obolibrary.org/obo/MEBDO_'+ zipId +'"/>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    print i
    updateFile.close()

updateFile = open('mebdo_DRG_hospital_individual_import.owl','a')
updateFile.write('</rdf:RDF>\n\n\n<!-- Generated by the python code -->\n')
updateFile.close()
