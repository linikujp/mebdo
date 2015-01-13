""" outpatient individual payment
   <!-- http://purl.obolibrary.org/obo/MEBDO_0000415 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000415">
        <rdf:type rdf:resource="&obo;MEBDO_0000043"/>
        <rdfs:label xml:lang="en">outpatient service 1</rdfs:label>
        <obo:MEBDO_0000026 rdf:resource="&obo;MEBDO_0000408"/>
        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_0000416"/>
        <obo:MEBDO_0000417 rdf:resource="&obo;MEBDO_0000418"/>
        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_0000419"/>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000416 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000416">
        <rdf:type rdf:resource="&obo;MEBDO_0000032"/>
        <rdfs:label xml:lang="en">161.58</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000418 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000418">
        <rdf:type rdf:resource="&obo;MEBDO_0000027"/>
        <rdfs:label xml:lang="en">1053</rdfs:label>
        <obo:MEBDO_0000028 rdf:datatype="&xsd;integer">2012</obo:MEBDO_0000028>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000419 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000419">
        <rdf:type rdf:resource="&obo;MEBDO_0000022"/>
        <rdfs:label xml:lang="en">33.59</rdfs:label>
    </owl:NamedIndividual>
    

"""
##hospital id will use 8+their id in the table. (ignor the reference region for the moment)
## asign the column I (type:MEBDO_0000027), J(type:MEBDO_0000032), K(type: MEBDO_0000408) first, then assign the hospital (column C)
## id starts from 9000000


from openpyxl import load_workbook

wb = load_workbook(filename = 'Medicare_Provider_Charge_Outpatient_APC30_CY2012.xlsx')
ws = wb.get_sheet_by_name( name = 'opps_apc_summary')

varId = 9000000

for i in range(8, 509):
    varId = 1 + varId 
    apc = ws.cell('A'+ str(i)).value
    hospitalId = ws.cell('B'+ str(i)).value
    hospitalId = '8' + str(hospitalId)
    serviceCount = ws.cell('I'+ str(i)).value
    charge = ws.cell('J'+ str(i)).value
    payment = ws.cell('K'+ str(i)).value
    ## get the APC id : split the column A and get the numbers before , then add 272
    apc = str(apc)
    apcCode = apc.split("-",2)[0]
    apcId = 'MEBDO_272' + apcCode.rstrip()
    
    updateFile = open('mebdo_APC_payment_individual.owl','a')
    
    serviceCountId = str(varId)
    serviceCount = str(serviceCount)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + serviceCountId + ' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_'+ serviceCountId +'">\n')
    updateFile.write('        <rdf:type rdf:resource="&obo;MEBDO_0000027"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">' + serviceCount + '</rdfs:label>\n')
    updateFile.write('        <obo:MEBDO_0000028 rdf:datatype="&xsd;integer">2012</obo:MEBDO_0000028>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    
    varId = varId + 1
    chargeId = str(varId)
    charge = str(charge)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_'+ chargeId +' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_'+ chargeId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="&obo;MEBDO_0000032"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ charge +'</rdfs:label>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n') 

    varId = varId + 1
    paymentId = str(varId)
    payment = str(payment)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_'+ paymentId +' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_'+ paymentId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="&obo;MEBDO_0000408"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+ payment +'</rdfs:label>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n') 
    
    varId = varId + 1
    serviceId = str(varId)
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_'+ serviceId +' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MEBDO_'+ serviceId + '">\n')
    updateFile.write('        <rdf:type rdf:resource="&obo;'+ apcId +'"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">outpatient service '+serviceId+'</rdfs:label>\n')
    updateFile.write('        <obo:MEBDO_0000026 rdf:resource="&obo;MEBDO_' + hospitalId + '"/>\n')
    updateFile.write('        <obo:MEBDO_0000417 rdf:resource="&obo;MEBDO_'+ serviceCountId +'"/>\n')
    updateFile.write('        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_' + paymentId + '"/>\n')
    updateFile.write('        <obo:MEBDO_0000414 rdf:resource="&obo;MEBDO_' + chargeId + '"/>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    
    print i
    updateFile.close()