"""
    <!-- http://purl.obolibrary.org/obo/MEBDO_0000038 -->

    <owl:Class rdf:about="&obo;MEBDO_0000038">
        <rdfs:label xml:lang="en">HEART TRANSPLANT OR IMPLANT OF HEART ASSIST SYSTEM W MCC</rdfs:label>
        <rdfs:subClassOf rdf:resource="&obo;MEBDO_0000029"/>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;BFO_0000117"/>
                <owl:someValuesFrom rdf:resource="&obo;MEBDO_0000035"/>
            </owl:Restriction>
        </rdfs:subClassOf>
        <obo:MEBDO_0000039 rdf:datatype="&xsd;integer">001</obo:MEBDO_0000039>
    </owl:Class>

"""

## The DRG group start with 374, and make it to 7digits. We add 374 0[4-n] code ; n = length of the string

from openpyxl import load_workbook

wb = load_workbook(filename = 'IPPS_FY2012MSDRG.xlsx')

ws = wb.get_sheet_by_name( name = 'Table1')

for i in range(5, 756):
    MSDRGcode = ws.cell('A'+ str(i)).value
    MSDRGcode = str(MSDRGcode)
    n = len(MSDRGcode.strip())
    string_val = "0" * (4-n)
    string_MSDRG = "0" * (3-n) + MSDRGcode.strip()
    indexString = '374' + string_val + MSDRGcode.strip()
    
    className = ws.cell('D'+ str(i)).value
    className = str(className)
    
    type = ws.cell('C'+ str(i)).value
    type = str(type)

    updateFile = open('mebdo_MSDRG.owl','a')

    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + indexString)
    updateFile.write(' -->')
    updateFile.write('\n')
    updateFile.write('\n')
    updateFile.write('    <owl:Class rdf:about="&obo;MEBDO_' + indexString)
    updateFile.write('">\n')
    
    updateFile.write('        <rdfs:label xml:lang="en">' + className.strip())
    updateFile.write('</rdfs:label>\n')
    updateFile.write('        <rdfs:subClassOf rdf:resource="&obo;MEBDO_0000029"/>\n')

    
    if type.strip() == "SURG":
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;BFO_0000117"/>\n')
        updateFile.write('                <owl:someValuesFrom rdf:resource="&obo;MEBDO_0000035"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')

    if type.strip() == "MED":
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;BFO_0000117"/>\n')
        updateFile.write('                <owl:someValuesFrom rdf:resource="&obo;MEBDO_0000036"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')

    updateFile.write('        <obo:MEBDO_0000039 rdf:datatype="&xsd;integer">'+ string_MSDRG+'</obo:MEBDO_0000039>\n    </owl:Class>\n\n\n')
    print i
    updateFile.close()